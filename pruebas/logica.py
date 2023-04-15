import openpyxl

libro = openpyxl.load_workbook('Libro2.xlsx')
hojas = libro.worksheets #crear la lista de las hojas de mi archivo
hoja = hojas[0] #indicando la primer hoja 


clientes ={}#creando el diccionario donde se almacenaran los clientes y los saldos totales


for fila in range(2, hoja.max_row +1): #iterar sobre la base de datos para llenar los datos del diccionario de clientes

    cliente = hoja.cell(row=fila, column=1).value #seleccionar los valores de la columna de clientes 
    varlor = hoja.cell(row=fila, column=3).value #seleccionar los valores  de mi lista de valores

    clientes.setdefault(cliente, 0) #ingresar las keys de mi diccionario de clientes
    clientes[cliente] += varlor #sumar los valores encontrados de cada cliente e ingresarlo en el diccionario



def create_customer_table():

    #crear el titulo de la tabla donde se va a comparar
    tabla_clientes = hoja.cell(row=1, column= hoja.max_column +3, value='Clientes') #crear el encabezado de clientes para la tabla que va a tener los valores a comparar
    tabla_total = hoja.cell(row=1, column= hoja.max_column +1, value='Total') 
    tabla_total_declarado = hoja.cell(row=1, column= hoja.max_column +1, value='Total declarado')
    igualar = hoja.cell(row=1, column= hoja.max_column +1, value='Igualar')
    columna = hoja.max_column - 3 #seleccionando la columna doonde se ingresan los nombres de los clientes
    num = 2 #creando una variable para ir aumentando el

    for clave in clientes.keys():
        persona = str(clave)
        cliente = hoja.cell(row=num, column=columna)
        cliente.value = persona
        # print(num)
        num = num + 1
        

    for fila in range(2, tabla_clientes):
        cliente = hoja.cell(row=fila, column=columna).value
        total = hoja.cell(row=fila, column=columna+1)

        total.value = clientes[cliente]
        print(total.value)

create_customer_table()

# prueba = hoja.cell(row=5, column=6).value

# print(prueba)
