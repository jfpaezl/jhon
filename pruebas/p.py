import openpyxl

libro = openpyxl.load_workbook('Libro2.xlsx')
hojas = libro.worksheets #crear la lista de las hojas de mi archivo
hoja = hojas[0] #indicando la primer hoja 
hoja_reporte = libro.create_sheet('Reporte')

clientes ={}#creando el diccionario donde se almacenaran los clientes y los saldos totales

tabla_reporte = [
    'Clientes', 
    'Total',
    'Total declarado',
    'Igualar'
    ]

def buscar_columna(fila : int, encabezado : str):
    # Identificar en qué fila se encuentra el encabezado
    fila_encabezado = fila  # Por ejemplo, si el encabezado está en la primera fila

    # Recorrer las celdas de la fila buscando el encabezado con el nombre específico
    encabezado_buscado = encabezado  # Por ejemplo, buscamos el encabezado 'Nombre'
    columna_encabezado = None
    for columna in range(1, hoja_reporte.max_column + 1):
        celda = hoja_reporte.cell(row=fila_encabezado, column=columna)
        if celda.value == encabezado_buscado:
            columna_encabezado = columna
            break
    
    # # Verificar si se encontró el encabezado y en qué columna se encuentra
    # if columna_encabezado is None:
    #     print(f"No se encontró el encabezado '{encabezado_buscado}' en la hoja '{hoja.title}'")
    # else:
    #     print(f"El encabezado '{encabezado_buscado}' se encuentra en la columna {columna_encabezado}")

def llenar_diccionario():

    for fila in range(2, hoja.max_row +1): #iterar sobre la base de datos para llenar los datos del diccionario de clientes

        cliente = hoja.cell(row=fila, column=1).value #seleccionar los valores de la columna de clientes 
        varlor = hoja.cell(row=fila, column=3).value #seleccionar los valores  de mi lista de valores

        clientes.setdefault(cliente, 0) #ingresar las keys de mi diccionario de clientes
        clientes[cliente] += varlor #sumar los valores encontrados de cada cliente e ingresarlo en el diccionario



def crear_tabla_clientes():

    #crear el titulo de la tabla donde se va a comparar
    hoja_reporte.cell(row=1, column= hoja_reporte.max_column, value='Clientes') #crear el encabezado de clientes para la tabla que va a tener los valores a comparar
    hoja_reporte.cell(row=1, column= hoja_reporte.max_column +1, value='Total') 
    hoja_reporte.cell(row=1, column= hoja_reporte.max_column +1, value='Total declarado')
    hoja_reporte.cell(row=1, column= hoja_reporte.max_column +1, value='Igualar')

    # Identificar en qué fila se encuentra el encabezado
    fila_encabezado = 1  # Por ejemplo, si el encabezado está en la primera fila

    # Recorrer las celdas de la fila buscando el encabezado con el nombre específico
    encabezado_buscado = 'Cliente'  # Por ejemplo, buscamos el encabezado 'Nombre'
    columna_encabezado = None
    for columna in range(1, hoja_reporte.max_column + 1):
        celda = hoja_reporte.cell(row=fila_encabezado, column=columna)
        if celda.value == encabezado_buscado:
            columna_encabezado = columna
            break



    columna = hoja_reporte.max_column #seleccionando la columna doonde se ingresan los nombres de los clientes
    num = 2 #creando una variable para ir aumentando el

    for clave in clientes.keys():
        persona = str(clave)
        cliente = hoja_reporte.cell(row=num, column=columna)
        cliente.value = persona
        # print(num)
        num = num + 1
        
    columna = encabezado_buscado(1, tabla_reporte[0])
    
    for fila in range(2, columna.max_row):
        cliente = hoja.cell(row=fila, column=columna).value
        total = hoja.cell(row=fila, column=columna+1)

        total.value = clientes[cliente]
        print(total.value)

"""
llamar las funciones 
"""
llenar_diccionario()
# crear_tabla_clientes()

prueba = hoja_reporte.cell(row=2, column=1).value

print(clientes)
