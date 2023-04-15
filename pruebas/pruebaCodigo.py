import openpyxl

# Cargar el archivo de Excel
libro = openpyxl.load_workbook('Libro2.xlsx')

# Seleccionar la hoja en la que se encuentra el encabezado
hojas = libro.worksheets #crear la lista de las hojas de mi archivo
hoja = hojas[0] #indicando la primer hoja 
hoja_reporte = hojas[1]

def prueba_buscar_titulo():

    # Identificar en qué fila se encuentra el encabezado
    fila_encabezado = 1  # Por ejemplo, si el encabezado está en la primera fila

    # Recorrer las celdas de la fila buscando el encabezado con el nombre específico
    encabezado_buscado = 'valor'  # Por ejemplo, buscamos el encabezado 'Nombre'
    columna_encabezado = None
    for columna in range(1, hoja.max_column + 1):
        celda = hoja.cell(row=fila_encabezado, column=columna)
        if celda.value == encabezado_buscado:
            columna_encabezado = columna
            break

    # Verificar si se encontró el encabezado y en qué columna se encuentra
    if columna_encabezado is None:
        print(f"No se encontró el encabezado '{encabezado_buscado}' en la hoja '{hoja.title}'")
    else:
        print(f"El encabezado '{encabezado_buscado}' se encuentra en la columna {columna_encabezado}")

def prueba_buscar_columna_porCelda():

    # Obtener la celda seleccionada
    celda_seleccionada = hoja['B1']

    # Obtener la columna de la celda seleccionada
    columna = celda_seleccionada.column

    print("La celda seleccionada se encuentra en la columna", columna)

def crear_encabezado():
    columna_encabezado = hoja_reporte.max_column

    