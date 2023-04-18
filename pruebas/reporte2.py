import openpyxl
import datetime

LISTA_CLIENTES = "lista_clientes"
NOMBRE_CLIENTE = "nombre"
VALOR_CLIENTE = "valor"
VALOR_DECLARADO = "declaracion"
COMPARACION = "comparacion"
TOTAL_LABEL = "Total"

"""variable creada para colocar el nombre del archivo el cual se desea hacer el reporte"""
nombre_archivo = "Libro1" 
nombre_archivo_final = 'Reporte'

libro = openpyxl.load_workbook(f'{nombre_archivo}.xlsx')
hojas = libro.worksheets #crear la lista de las hojas de mi archivo
hoja = hojas[0] #indicando la primer hoja 
hoja_reporte = libro.create_sheet('Reporte')
total = 0
cliente = {}

"""creando el diccionario donde se almacenaran los clientes y los saldos totales"""
clientes ={LISTA_CLIENTES: []}

"""lista para crear la tabla del reporte"""
lista_reporte = [ 
    'Clientes', 
    'Total',
    'Total declarado',
    'Comparación'
    ] 


"""iterar sobre la tabla para llenar el diccionario de clientes con sus totales a pagar"""

for fila in range(2, hoja.max_row +1): #iterar sobre la base de datos para llenar los datos del diccionario de clientes
    concepto = hoja.cell(row=fila, column=1).value #seleccionar los valores de la columna de clientes 
    valor = hoja.cell(row=fila, column=2).value #seleccionar los valores  de mi lista de valores
    if valor and concepto != TOTAL_LABEL:
        total += valor
        
    if not valor:
        
        cliente.setdefault(NOMBRE_CLIENTE, concepto)
    if concepto == TOTAL_LABEL:
        cliente.setdefault(VALOR_CLIENTE, total) 
        cliente.setdefault(VALOR_DECLARADO, valor)
        cliente.setdefault(COMPARACION, valor-total)
        clientes[LISTA_CLIENTES].append(cliente)
        total = 0
        cliente = {}
""" buscar la columna de un encabezado"""
def buscar_columna(fila : int, encabezado : str, hoja):
    # Recorrer las celdas de la fila buscando el encabezado con el nombre específico
    encabezado_buscado = encabezado  # Por ejemplo, buscamos el encabezado 'Cliente'
    columna_encabezado = None
    for columna in range(1, hoja.max_column + 1):
        celda = hoja.cell(row=fila, column=columna)
        if celda.value == encabezado_buscado:
            columna_encabezado = columna
            break
    return columna_encabezado

"""crear los encabezados de la hoja de reporte"""
def crear_encabezado(hoja : str, lista : list):
    #crear el titulo de la lista donde se va a comparar
    columna = hoja.max_column
    for encabezado in lista:
        # print(encabezado)
        hoja.cell(row=1, column= columna, value=encabezado)
        columna+=1
    
    # print(hoja.cell(row=1, column=4).value)
def rellnar_tabla(hoja_reporte, lista_reporte):
    crear_encabezado(hoja_reporte, lista_reporte)
    columna_clientes = buscar_columna(1, 'Clientes', hoja_reporte)
    columna_total = buscar_columna(1, 'Total', hoja_reporte)
    columna_declarado = buscar_columna(1, 'Total declarado', hoja_reporte)
    columna_comparacion = buscar_columna(1, 'Comparación', hoja_reporte)
    num = 2 #creando una variable para ir aumentando el
    # print(columna_clientes)
    """rellenar la columna de clientes en la hoja de Reporte"""

    for cliente in clientes[LISTA_CLIENTES]:
        variable = cliente[NOMBRE_CLIENTE]
        cliente = hoja_reporte.cell(row=num, column=columna_clientes) #seleccionar la casilla que se desea rellenar
        cliente.value = variable
        num += 1
    num = 2
    for cliente in clientes[LISTA_CLIENTES]:
        variable = cliente[VALOR_CLIENTE]
        cliente = hoja_reporte.cell(row=num, column=columna_total) #seleccionar la casilla que se desea rellenar
        cliente.value = variable
        num += 1
    num = 2
    for cliente in clientes[LISTA_CLIENTES]:
        variable = cliente[VALOR_DECLARADO]
        cliente = hoja_reporte.cell(row=num, column=columna_declarado) #seleccionar la casilla que se desea rellenar
        cliente.value = variable
        # print(cliente)
        num += 1
    num = 2
    for cliente in clientes[LISTA_CLIENTES]:
        variable = cliente[COMPARACION]
        cliente = hoja_reporte.cell(row=num, column=columna_comparacion) #seleccionar la casilla que se desea rellenar
        cliente.value = variable
        # print(cliente)
        num += 1

"""fecha del reporte"""
rellnar_tabla(hoja_reporte, lista_reporte)

"""fecha del reporte"""
fecha_actual = datetime.date.today()

libro.save(f'{nombre_archivo_final} {fecha_actual}.xlsx')