from funciones.funciones import * #importar primero carpeta luego con el . se selecciona el archivo y el import es para extraer las funciones deseadas
import openpyxl
import datetime

"""variable creada para colocar el nomvre del archivo el cual se desea hacer el reporte"""
nombre_archivo = "Libro2" 
nombre_archivo_final = 'Reporte'

libro = openpyxl.load_workbook(f'{nombre_archivo}.xlsx')
hojas = libro.worksheets #crear la lista de las hojas de mi archivo
hoja = hojas[0] #indicando la primer hoja 
hoja_reporte = libro.create_sheet('Reporte')

"""creando el diccionario donde se almacenaran los clientes y los saldos totales"""
clientes ={}

"""lista para crear la rapla del reporte"""
lista_reporte = [ 
    'Clientes', 
    'Total',
    'Total declarado',
    'Igualar'
    ] 

"""iterar sobre la tabla para llenar el diccionario de clientes con sus totales a pagar"""
for fila in range(2, hoja.max_row +1): #iterar sobre la base de datos para llenar los datos del diccionario de clientes
    cliente = hoja.cell(row=fila, column=1).value #seleccionar los valores de la columna de clientes 
    valor = hoja.cell(row=fila, column=3).value #seleccionar los valores  de mi lista de valores
    clientes.setdefault(cliente, 0) #ingresar las keys de mi diccionario de clientes
    

    clientes[cliente] += valor #sumar los valores encontrados de cada cliente e ingresarlo en el diccionario

"""llamar la funcion para rellenar la tabla"""
rellenar_tabla(hoja_reporte, lista_reporte, clientes)

# prueba = hoja_reporte.cell(row=2, column=2).value
# print(prueba)

"""fecha del reporte"""
fecha_actual = datetime.date.today()

libro.save(f'{nombre_archivo_final} {fecha_actual}.xlsx')